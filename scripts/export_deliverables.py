# -*- coding: utf-8 -*-
"""
export_deliverables.py · 交付层导出(V71 双产出层)

职责:把工具链内部产出(output/ / c_mode/ / b_mode/ / final_tender_package/)
     按 DELIVERABLE_MAPPING 映射到中文命名的 Office 格式交付层
     (投标交付物/),供投标当天直接使用。

核心原则:
- 交付层独立于工具链内部(business_model §8 #N24)
- 每次执行覆盖上次产出(幂等)
- 交付层不进 tracked_outputs,不参与 baseline
- 用户手工修改交付层不污染工具链

用法:
    python scripts/export_deliverables.py --project demo_cadre_training
    python scripts/export_deliverables.py --project demo_cadre_training --dry-run
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[错误] 缺少 openpyxl 依赖。pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent


# ──────────────────────────────────────────────────────────
# 映射表(v2 单元 7 P19:从 response_file_parts 动态构建)
# ──────────────────────────────────────────────────────────

# 保留一组"顶层固定映射",针对所有项目通用的交付文件
STATIC_TOP_MAPPING = [
    {
        "source": "final_tender_package/final_response.docx",
        "target": "最终投标文件.docx",
        "mode": "copy",
    },
    {
        "source": "output/scoring_matrix.csv",
        "target": "评分矩阵.xlsx",
        "mode": "csv_to_xlsx",
    },
]


def _safe_part_dir_name(name: str) -> str:
    """与 c_mode_extract / b_mode_extract 的 _safe_name 保持一致的目录命名规则。"""
    import re
    s = re.sub(r'^[一二三四五六七八九十\d]+[、.]\s*', '', name)
    s = re.sub(r'[(].+?[)]', '', s)
    s = re.sub(r'[(].+?[)]', '', s)
    return s.strip() or 'part'


def build_deliverable_mapping(brief: dict) -> list[dict]:
    """
    v2 单元 7 P19:按 tender_brief.response_file_parts 动态生成映射表。
    不再硬编码另一项目的 Part 名。
    """
    mapping = list(STATIC_TOP_MAPPING)  # copy
    parts = brief.get('response_file_parts', [])

    for i, part in enumerate(parts):
        mode = (part.get('production_mode') or '').strip().upper()
        sub = (part.get('sub_mode') or '').strip()
        name = part.get('name', f'Part_{i}')
        dir_name = _safe_part_dir_name(name)
        order = part.get('order', i + 1)

        if mode == 'C':
            if sub == 'C-template':
                mapping.append({
                    "source": f"output/c_mode/{dir_name}/filled.docx",
                    "target": f"C 模式产出/{order:02d}_{name}.docx",
                    "mode": "copy",
                })
            elif sub == 'C-reference':
                mapping.append({
                    "source": f"output/c_mode/{dir_name}/instructions.md",
                    "target": f"C 模式产出/{order:02d}_{name}操作说明.md",
                    "mode": "copy",
                })
            elif sub == 'C-attachment':
                # V4-4: 映射 attachment 目录 + 同位置带 _manifest.yaml (占位状态透传)。
                # source dir 是 v45_merge 拷贝过去的 final_tender_package/attachments/<part>/,
                # 只含 status=resolved 的真实附件(pending_user 项不在物理目录里, 不假装存在)。
                # _manifest.yaml 同步导出, 让 bidder 在交付层直接看到各附件 status。
                mapping.append({
                    "source": f"final_tender_package/attachments/{dir_name}",
                    "target": f"C 模式产出/{order:02d}_{name}(附件)",
                    "mode": "copy_dir",
                })
                mapping.append({
                    "source": f"output/c_mode/{dir_name}/attachments.yaml",
                    "target": f"C 模式产出/{order:02d}_{name}(附件)/_manifest.yaml",
                    "mode": "copy",
                })
        elif mode == 'B':
            mapping.append({
                "source": f"output/b_mode/{dir_name}/assembled.docx",
                "target": f"B 模式产出/{order:02d}_{name}(占位).docx",
                "mode": "copy",
            })
        # A 模式已通过 final_response.docx 合并,单独映射"技术部分分章节"供核对
        elif mode == 'A':
            mapping.append({
                "source": "output/tender_response.docx",
                "target": f"A 模式产出/{order:02d}_{name}(分章节).docx",
                "mode": "copy",
            })
        # D / 不适用 → 不映射

    return mapping


# ──────────────────────────────────────────────────────────
# AI prompt: md → xlsx 结构化转换
# ──────────────────────────────────────────────────────────

# v1 V77 清理: MD_TO_XLSX_PROMPT + 6 个启发式 md 解析函数已删除
# (_split_md_by_section, _extract_field_inline, _extract_bullet_list_after_label,
#  _parse_operations_checklist, _parse_pending_manual_work, parse_md_to_rows)
# 原因: 违反 skill 哲学三条红线(对非结构化 md 做语义切分/字段映射)。
# 新流程: md → xlsx 由主 agent 在对话里组 rows dict 列表,然后调
# _write_xlsx_from_rows 写入。详见 SKILL.md 阶段 6。


def _do_copy(src: Path, dst: Path) -> dict:
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    return {"mode": "copy", "src_size": src.stat().st_size, "dst_size": dst.stat().st_size}


def _do_copy_dir(src: Path, dst: Path) -> dict:
    """V4-4: 目录级拷贝 (C-attachment 用)。dirs_exist_ok=True 允许同目标多次写入
    (如 attachments 目录 + 后续追加的 _manifest.yaml 共存)。
    """
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copytree(src, dst, dirs_exist_ok=True)
    n_files = sum(1 for _ in dst.rglob('*') if _.is_file())
    total_size = sum(p.stat().st_size for p in dst.rglob('*') if p.is_file())
    return {"mode": "copy_dir", "n_files": n_files, "dst_size": total_size}


def _apply_xlsx_style(ws, n_cols: int):
    """首行加粗 + 冻结首行 + 列宽自适应 + 自动换行。"""
    bold_font = Font(bold=True)
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'
    # 列宽自适应(按单元格内容长度估算)
    for col_idx in range(1, n_cols + 1):
        max_len = 0
        for cell in ws[get_column_letter(col_idx)]:
            val = cell.value
            if val is None:
                continue
            s = str(val)
            # 每行取最长一段
            for ln in s.split('\n'):
                # 中文字符算 2
                width = sum(2 if ord(c) > 127 else 1 for c in ln)
                if width > max_len:
                    max_len = width
        # 最小 10, 最大 60
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)
    # 所有数据行单元格自动换行
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')


def _do_csv_to_xlsx(src: Path, dst: Path) -> dict:
    """csv → xlsx,首行加粗+冻结+列宽自适应。"""
    dst.parent.mkdir(parents=True, exist_ok=True)
    with open(src, 'r', encoding='utf-8-sig', newline='') as f:
        rows = list(csv.reader(f))
    if not rows:
        raise ValueError(f"CSV 为空: {src}")

    wb = Workbook()
    ws = wb.active
    ws.title = "scoring_matrix"
    for row in rows:
        ws.append(row)
    _apply_xlsx_style(ws, len(rows[0]))
    wb.save(str(dst))
    return {
        "mode": "csv_to_xlsx",
        "rows": len(rows) - 1,
        "cols": len(rows[0]),
        "dst_size": dst.stat().st_size,
    }


def _write_xlsx_from_rows(dst: Path, headers: list[str], rows: list[dict],
                          sheet_title: str = 'Sheet1') -> dict:
    """v1 V77: 纯 openpyxl 写入。

    主 agent 在对话里读 md 内容、按 schema 组织 rows(list of dict),
    然后调本函数完成确定性写入。脚本不解析 md(语义切分属红线),
    主 agent 做语义结构化。详见 SKILL.md 阶段 6 + business_model §9。

    参数:
        dst: 目标 xlsx 路径
        headers: 列名列表(同时作为字典 key)
        rows: 数据行,每项是 {字段名: 值} dict
        sheet_title: 工作表名,可选

    返回:写入统计 dict。
    """
    dst.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    for row_dict in rows:
        ws.append([row_dict.get(col, '') for col in headers])
    _apply_xlsx_style(ws, len(headers))
    wb.save(str(dst))
    return {
        "mode": "xlsx_from_rows",
        "rows": len(rows),
        "cols": len(headers),
        "dst_size": dst.stat().st_size,
    }


# ──────────────────────────────────────────────────────────
# 主流程
# ──────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='交付层导出(V71)')
    parser.add_argument('--project', required=True, help='项目名')
    parser.add_argument('--dry-run', action='store_true',
                        help='仅打印映射表,不实际导出')
    args = parser.parse_args()

    # v3.0.1 路径穿越防护: project 必须是单个目录名(脚本会拼到
    # ROOT/projects/<args.project>),不允许含 .. 或路径分隔符,
    # 否则可能逃逸到 projects/ 之外的目录
    if ".." in args.project or "/" in args.project or "\\" in args.project:
        print(
            f"[错误] --project 必须是单个目录名,不允许含 '..' 或路径分隔符 '/' '\\\\'。"
            f"got {args.project!r}",
            file=sys.stderr,
        )
        sys.exit(1)

    project_dir = ROOT / 'projects' / args.project

    # V3-3: timing hook
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from _timing_hook import stage_timer

    with stage_timer("export_deliverables", project_dir):
        _main_body(args, project_dir)


def _main_body(args, project_dir):
    if not project_dir.exists():
        print(f"[错误] 项目目录不存在: {project_dir}", file=sys.stderr)
        sys.exit(1)

    deliverable_root = project_dir / '投标交付物'

    # v2 单元 7 P19:从 tender_brief.json 动态构建映射表
    brief_path = project_dir / 'output' / 'tender_brief.json'
    if not brief_path.exists():
        print(f"[错误] 找不到 {brief_path},先跑阶段 1", file=sys.stderr)
        sys.exit(1)
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from brief_schema import load_brief_guarded
    brief = load_brief_guarded(brief_path)
    mapping = build_deliverable_mapping(brief)

    if args.dry_run:
        print("=" * 60)
        print(f"[dry-run] 映射表(共 {len(mapping)} 项):")
        print("=" * 60)
        for i, item in enumerate(mapping, 1):
            print(f"{i:2}. [{item['mode']:14s}] {item['source']}")
            print(f"         → 投标交付物/{item['target']}")
            if 'schema' in item:
                print(f"         schema: {item['schema']}")
        return

    # 清空交付层目录(每次投标重新生成)
    if deliverable_root.exists():
        shutil.rmtree(deliverable_root)
    deliverable_root.mkdir(parents=True, exist_ok=True)

    print(f"[信息] 导出到: {deliverable_root}")
    print(f"[信息] 映射表项数: {len(mapping)}")
    print()

    stats_per_mode = {"copy": 0, "csv_to_xlsx": 0, "copy_dir": 0}
    missing = []
    for item in mapping:
        src = project_dir / item['source']
        if not src.exists():
            missing.append(item['source'])
            continue
    if missing:
        print(f"[错误] 以下源文件缺失,无法导出:", file=sys.stderr)
        for m in missing:
            print(f"  - {m}", file=sys.stderr)
        sys.exit(1)

    for i, item in enumerate(mapping, 1):
        src = project_dir / item['source']
        dst = deliverable_root / item['target']
        mode = item['mode']
        try:
            if mode == 'copy':
                info = _do_copy(src, dst)
            elif mode == 'csv_to_xlsx':
                info = _do_csv_to_xlsx(src, dst)
            elif mode == 'copy_dir':
                info = _do_copy_dir(src, dst)
            else:
                raise ValueError(
                    f"未知 mode: {mode}。v1 后 md→xlsx 不再走映射表,"
                    f"由主 agent 调 _write_xlsx_from_rows 处理。"
                )
        except Exception as e:
            print(f"[错误] 导出 {item['target']} 失败: {e}", file=sys.stderr)
            sys.exit(1)

        stats_per_mode[mode] += 1
        extra = ''
        if mode == 'csv_to_xlsx':
            extra = f" ({info['rows']} rows × {info['cols']} cols)"
        print(f"  [{i:2}/{len(mapping)}] [{mode:14s}] "
              f"{item['target']} ({info.get('dst_size', 0):,} bytes){extra}")

    print()
    print("=" * 60)
    print(f"[完成] 交付层导出完成 → {deliverable_root}")
    print(f"  copy:         {stats_per_mode['copy']} 项")
    print(f"  copy_dir:     {stats_per_mode['copy_dir']} 项(V4-4 C-attachment)")
    print(f"  csv_to_xlsx:  {stats_per_mode['csv_to_xlsx']} 项")
    print(f"  md_to_xlsx:   主 agent 任务,不在自动映射(SKILL.md 阶段 6)")
    print()
    print(f"  目录结构:")
    for sub in sorted(deliverable_root.rglob('*')):
        if sub.is_file():
            rel = sub.relative_to(deliverable_root)
            print(f"    投标交付物/{rel} ({sub.stat().st_size:,} bytes)")
    print("=" * 60)


if __name__ == '__main__':
    main()
